"""Parsing and validation of client store files (.xlsx / .csv).

Turns an uploaded "BDD client" file into an import preview: the list of valid
stores, plus one explicit error per invalid cell (with the file row number) so
the frontend can show exactly what to fix. Nothing is persisted at this stage.

CSV is read with the standard library (UTF-8 or Windows-1252, comma or
semicolon separated — French Excel exports use semicolons). Excel files are
read with ``openpyxl``.
"""
from __future__ import annotations

import csv
import io
import re
from pathlib import Path

#: Columns the client file must contain (header names, case-insensitive).
REQUIRED_COLUMNS = [
    "store_id",
    "name",
    "city",
    "address",
    "latitude",
    "longitude",
    "opening_hours",
    "store_url",
]

#: Text fields that only need to be non-empty.
_TEXT_FIELDS = ("store_id", "name", "city", "address", "opening_hours")

#: Loose-but-useful URL shape: http(s)://… or www.… with at least one dot.
_URL_RE = re.compile(r"^(https?://|www\.)\S+\.\S{2,}$", re.IGNORECASE)

#: Safety cap on uploads (the client DB files stay small).
_MAX_SIZE_BYTES = 5 * 1024 * 1024


class StoreImportError(ValueError):
    """File-level problem (unsupported format, unreadable, empty…)."""


# --------------------------------------------------------------------------- #
#  Cell / grid helpers                                                         #
# --------------------------------------------------------------------------- #
def _cell(value) -> str:
    """Normalize any cell value to a stripped string ('' for None)."""
    if value is None:
        return ""
    # openpyxl gives floats for numeric cells; keep "101" instead of "101.0".
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _to_float(raw: str) -> float | None:
    """Parse a decimal that may use a comma separator ('33,58')."""
    try:
        return float(raw.replace(" ", "").replace(",", "."))
    except ValueError:
        return None


def _read_csv(content: bytes) -> list[list[str]]:
    for encoding in ("utf-8-sig", "cp1252"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise StoreImportError(
            "Encodage du fichier non reconnu (enregistrez-le en UTF-8)."
        )
    try:
        dialect = csv.Sniffer().sniff(text[:2048], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel  # default: comma-separated
    return [[_cell(value) for value in row] for row in csv.reader(io.StringIO(text), dialect)]


def _read_xlsx(content: bytes) -> list[list[str]]:
    from openpyxl import load_workbook  # imported lazily (heavy dependency)

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # corrupt / password-protected / not a real xlsx
        raise StoreImportError(
            "Impossible de lire le fichier Excel (fichier corrompu ou protégé)."
        ) from exc
    try:
        sheet = workbook.worksheets[0]
        return [
            [_cell(value) for value in row]
            for row in sheet.iter_rows(values_only=True)
        ]
    finally:
        workbook.close()


# --------------------------------------------------------------------------- #
#  Row validation                                                              #
# --------------------------------------------------------------------------- #
def _validate_row(row_number: int, values: dict[str, str], seen_ids: dict[str, int]):
    """Return (store dict | None, list of error dicts) for one data row."""
    errors: list[dict] = []

    def add_error(field: str, message: str) -> None:
        errors.append({"row": row_number, "field": field, "message": message})

    for field in _TEXT_FIELDS:
        if not values[field]:
            add_error(field, "Champ obligatoire manquant.")

    store_id = values["store_id"]
    if store_id:
        if store_id in seen_ids:
            add_error(
                "store_id",
                f"store_id « {store_id} » en double (déjà présent ligne {seen_ids[store_id]}).",
            )
        else:
            seen_ids[store_id] = row_number

    latitude = longitude = None
    if not values["latitude"]:
        add_error("latitude", "Champ obligatoire manquant.")
    else:
        latitude = _to_float(values["latitude"])
        if latitude is None:
            add_error("latitude", "Valeur numérique attendue.")
        elif not -90 <= latitude <= 90:
            add_error("latitude", "Doit être comprise entre -90 et 90.")

    if not values["longitude"]:
        add_error("longitude", "Champ obligatoire manquant.")
    else:
        longitude = _to_float(values["longitude"])
        if longitude is None:
            add_error("longitude", "Valeur numérique attendue.")
        elif not -180 <= longitude <= 180:
            add_error("longitude", "Doit être comprise entre -180 et 180.")

    url = values["store_url"]
    if not url:
        add_error("store_url", "Champ obligatoire manquant.")
    elif not _URL_RE.match(url):
        add_error("store_url", "URL invalide (attendu : http(s)://… ou www.…).")

    if errors:
        return None, errors

    return (
        {
            "store_id": store_id,
            "name": values["name"],
            "city": values["city"],
            "address": values["address"],
            "latitude": latitude,
            "longitude": longitude,
            "opening_hours": values["opening_hours"],
            "store_url": url,
        },
        [],
    )


# --------------------------------------------------------------------------- #
#  Entry point                                                                 #
# --------------------------------------------------------------------------- #
def build_preview(filename: str, content: bytes) -> dict:
    """Parse and validate an uploaded store file; return the preview payload."""
    extension = Path(filename).suffix.lower()
    if extension not in (".csv", ".xlsx"):
        raise StoreImportError(
            "Format non supporté : fournissez un fichier .xlsx ou .csv."
        )
    if not content:
        raise StoreImportError("Le fichier est vide.")
    if len(content) > _MAX_SIZE_BYTES:
        raise StoreImportError("Fichier trop volumineux (5 Mo maximum).")

    grid = _read_csv(content) if extension == ".csv" else _read_xlsx(content)

    # First non-empty row is the header.
    header_index = next(
        (i for i, row in enumerate(grid) if any(cell for cell in row)), None
    )
    if header_index is None:
        raise StoreImportError("Le fichier ne contient aucune donnée.")

    headers = [cell.lower() for cell in grid[header_index]]
    missing_columns = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing_columns:
        return {
            "filename": filename,
            "total_rows": 0,
            "valid_count": 0,
            "error_count": 0,
            "stores": [],
            "errors": [],
            "missing_columns": missing_columns,
            "message": (
                "Colonnes obligatoires manquantes : "
                + ", ".join(missing_columns)
                + ". Corrigez l'en-tête du fichier puis relancez l'analyse."
            ),
        }

    column_index = {col: headers.index(col) for col in REQUIRED_COLUMNS}
    stores: list[dict] = []
    errors: list[dict] = []
    seen_ids: dict[str, int] = {}
    total_rows = 0
    error_rows = 0

    for row_number, raw in enumerate(
        grid[header_index + 1 :], start=header_index + 2
    ):
        if not any(cell for cell in raw):
            continue  # skip fully empty rows
        total_rows += 1
        values = {
            col: (raw[column_index[col]] if column_index[col] < len(raw) else "")
            for col in REQUIRED_COLUMNS
        }
        store, row_errors = _validate_row(row_number, values, seen_ids)
        if store is not None:
            stores.append(store)
        else:
            error_rows += 1
            errors.extend(row_errors)

    if total_rows == 0:
        message = "Le fichier ne contient aucune ligne de données sous l'en-tête."
    elif error_rows == 0:
        message = f"{total_rows} ligne(s) analysée(s) : toutes valides. Prêt pour l'import."
    else:
        message = (
            f"{total_rows} ligne(s) analysée(s) : {len(stores)} valide(s), "
            f"{error_rows} en erreur. Corrigez les lignes signalées puis relancez l'analyse."
        )

    return {
        "filename": filename,
        "total_rows": total_rows,
        "valid_count": len(stores),
        "error_count": error_rows,
        "stores": stores,
        "errors": errors,
        "missing_columns": [],
        "message": message,
    }
